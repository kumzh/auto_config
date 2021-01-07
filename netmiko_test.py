from netmiko import ConnectHandler
from netmiko.huawei import HuaweiSSH, HuaweiVrpv8SSH, HuaweiTelnet
from netmiko.huawei import HuaweiSmartAXSSH
import openpyxl
import os
import logging
import time
#path为程序所在路径
path = os.path.abspath('.')

file_name = os.path.join(path,'Log','Auto_config_log.txt')
logging.basicConfig(filename= file_name,level=logging.DEBUG,format='%(asctime)s - %(levelname)s - %(message)s')


#提取Excel文件中的信息
def Read_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    #获取Excel的行列
    row = sheet.max_row
    column = sheet.max_column
    device_list = {}
    #由于第一行是表头，所以行从2开始；（行和列都是从1开始）
    for i in range(2,row+1):
        device_list['device{0}'.format(i-1)] = []
        for j in range(1,column+1):
            value = sheet.cell(row = i,column = j).value
            device_list['device{0}'.format(i-1)].append(value)
    return device_list

def Connect(device,command_file):
    F = True 
    str = '正在连接{0}\n'.format(device['host'])
    print(str)
    logging.info(str)
    net_connect = ConnectHandler(**device)
    net_connect.enable()  #输入启用模式
    com_path = os.path.join(path,'Command',command_file)
    #读取配置文件，配置文件存放在path目录下
    for i in open(com_path,'r'):
        cmd = i.replace('\n',' ')
        print('正在执行命令：',cmd)
        result = net_connect.send_command(cmd)
        print(result)
    net_connect.disconnect() 

def Huawei(ip):  
    huawei = {
        'device_type':'huawei',
        'host':ip,
        'username':'admin',
        'password':'HUAwei@123',
    }
    return huawei
def H3c(ip): #如果是display 华三需要刷入额外命令，改变屏幕显示长度
    h3c = {
        'device_type':'hp_comware',
        'host':ip,
        'username':'admin',
        'password':'H3c@96577',
    }
    return h3c

#获取Excel文件中放入配置文件名。如果没有配置文件，执行默认配置。
def Command_file(device_list,device_name):
    filename =device_list[device_name][2]
    if filename != None:
        return filename
    else:
        return 'command.txt'
def Init(ip_file):
    device_list = Read_excel(ip_file)
    for device_name in device_list.keys():
        try:
            if device_list[device_name][1] == 'huawei':
                device = Huawei(device_list[device_name][0])
                commandfile = Command_file(device_list,device_name)
                Connect(device,commandfile)
                str1 = device['host'] + '命令执行完毕！'
                logging.info(str1)
            elif device_list[device_name][1] == 'h3c': #注意华三设备在执行命令时不会一下显示所有内容。
                device = H3c(device_list[device_name][0])
                commandfile = Command_file(device_list,device_name)
                Connect(device,commandfile)
                str2 = device['host'] + '命令执行完毕！'
                logging.info(str2)
            else:
                str3 = device['host'] + '未定义设备！'
                print(str3)
                logging.info(str3)
                pass
        except Exception as e:
            str4 = '连接超时：{0}'.format(e)
            print(str4)
            logging.info(str4)
            pass 
        time.sleep(1)

if __name__ == '__main__':
    ip_path = os.path.join(path,'Host_info','ip_add.xlsx')
    Init(ip_path)
    
